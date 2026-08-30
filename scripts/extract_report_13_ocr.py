#!/usr/bin/env python3
"""
ДЕЛО ТЕХ — Отчёт №13: Движение по импорту
С OCR (Tesseract) для автоматического извлечения данных

Использование:
    python3 extract_report_13_ocr.py --start 20.07.2026 --end 10.08.2026 --output report.xlsx
"""

import asyncio
import argparse
import re
from playwright.async_api import async_playwright
from PIL import Image, ImageEnhance
import pytesseract
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Конфигурация
LOGIN = "pl_11640"
PASSWORD = "Qwerty123"
BASE_URL = "https://rlisystems.ru/conterra/"
MENU_DOP_USLUGI = (200, 570)
MENU_OTCHETNOST = (150, 700)


def preprocess_for_ocr(image_path):
    """Предобработка изображения для лучшего OCR"""
    img = Image.open(image_path)
    
    # Обрезаем только область таблицы
    width, height = img.size
    left = int(width * 0.15)
    top = int(height * 0.25)
    right = int(width * 0.98)
    bottom = int(height * 0.95)
    
    table_img = img.crop((left, top, right, bottom))
    
    # Увеличиваем размер (2x)
    table_img = table_img.resize((table_img.width * 2, table_img.height * 2), Image.LANCZOS)
    
    # Увеличиваем контраст
    enhancer = ImageEnhance.Contrast(table_img)
    table_img = enhancer.enhance(2.0)
    
    # Увеличиваем резкость
    enhancer = ImageEnhance.Sharpness(table_img)
    table_img = enhancer.enhance(2.0)
    
    return table_img


def parse_ocr_text(text):
    """Парсит текст OCR в структурированные данные"""
    lines = text.strip().split('\n')
    data = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Ищем номер контейнера (TKRU + 7 цифр)
        container_match = re.search(r'TKRU\d{7}', line)
        if container_match:
            container = container_match.group()
            
            row = {
                'container': container,
                'iso': '',
                'size': '',
                'capacity': '',
                'gross': '',
                'netto': '',
                'tare': '',
                'tare_max': '',
                'empty': '',
                'seals': '',
                'consignment': '',
                'empty_as_cargo': '',
                'cons_date': '',
                'cargo_name': '',
                'shipper': ''
            }
            
            # ISO тип
            iso_match = re.search(r'\b\d{2}[A-Z]\d\b', line)
            if iso_match:
                row['iso'] = iso_match.group()
            
            # Размер
            size_match = re.search(r'(DC|HC)\s*(20|40)', line)
            if size_match:
                row['size'] = f"{size_match.group(1)} {size_match.group(2)}"
            
            # Вместимость
            cap_match = re.search(r'\b(DC|HC)\b', line)
            if cap_match and not row['size']:
                row['capacity'] = cap_match.group(1)
            
            # Веса
            weights = re.findall(r'\d{2,5}\.\d{3}', line)
            if len(weights) >= 3:
                row['gross'] = weights[0]
                row['netto'] = weights[1]
                row['tare'] = weights[2]
            if len(weights) >= 4:
                row['tare_max'] = weights[3]
            
            # Дата
            date_match = re.search(r'\d{2}\.\d{2}\.\d{4}', line)
            if date_match:
                row['cons_date'] = date_match.group()
            
            # Коносамент
            cons_match = re.search(r'NNLTXG\d+', line)
            if cons_match:
                row['consignment'] = cons_match.group()
            
            # Груз
            if 'КРЕМНИЙ' in line:
                row['cargo_name'] = 'КРЕМНИЙ МЕТАЛЛ'
            elif 'ЭЛЕКТРОД' in line:
                row['cargo_name'] = 'ЭЛЕКТРОДЫ'
            elif 'АНОД' in line:
                row['cargo_name'] = 'АНОДНЫЕ БЛОКИ'
            
            data.append(row)
    
    return data


def create_excel(data, start_date, end_date, output_file):
    """Создать Excel файл"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Движение по импорту"
    
    # Заголовок
    ws.merge_cells('A1:P1')
    ws['A1'] = "ТЕРМИНАЛ ВРАНГЕЛЬ ООО"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:P2')
    ws['A2'] = "Наличие и движение импорта"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:P3')
    ws['A3'] = f"с {start_date} по {end_date}"
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:P4')
    ws['A4'] = 'Экспедитор : ООО "РУСАЛТРАНС"'
    ws['A4'].font = Font(size=10)
    ws['A4'].alignment = Alignment(horizontal='left')
    
    # Заголовки
    headers = ['N', 'Контейнер', 'ISO тип', 'Размер', 'Вместимость', 'Вес брутто', 
               'Вес нетто', 'Вес тары', 'Трафарет', 'Порожний', 'Пломбы', 
               '№ коносамента', 'Порожний как груз', 'Дата коносамента', 
               'Наименование груза', 'Грузоотправитель']
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.border = thin_border
    
    # Данные
    for row_idx, row_data in enumerate(data, 7):
        values = [
            row_idx - 6,
            row_data.get('container', ''),
            row_data.get('iso', ''),
            row_data.get('size', ''),
            row_data.get('capacity', ''),
            row_data.get('gross', ''),
            row_data.get('netto', ''),
            row_data.get('tare', ''),
            row_data.get('tare_max', ''),
            row_data.get('empty', ''),
            row_data.get('seals', ''),
            row_data.get('consignment', ''),
            row_data.get('empty_as_cargo', ''),
            row_data.get('cons_date', ''),
            row_data.get('cargo_name', ''),
            row_data.get('shipper', '')
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Ширина колонок
    for col, width in [('A', 5), ('B', 15), ('C', 10), ('D', 10), ('E', 12), 
                        ('F', 12), ('G', 12), ('H', 12), ('I', 15), ('J', 12), 
                        ('K', 12), ('L', 25), ('M', 15), ('N', 15), ('O', 12), 
                        ('P', 35), ('Q', 35)]:
        ws.column_dimensions[col].width = width
    
    wb.save(output_file)
    print(f"Excel сохранён: {output_file}")


async def extract_report(start_date: str, end_date: str, output_file: str):
    """Извлечь отчёт #13 с OCR"""
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-dev-shm-usage']
        )
        page = await browser.new_page(viewport={'width': 1280, 'height': 1024})
        
        print("[1/7] Открываю сайт...")
        await page.goto(BASE_URL, wait_until='domcontentloaded', timeout=60000)
        await asyncio.sleep(10)
        
        print("[2/7] Авторизация...")
        await page.evaluate(f"""() => {{
            var inputs = document.querySelectorAll('.v-textfield');
            if(inputs.length >= 2) {{
                inputs[0].value = '{LOGIN}';
                inputs[0].dispatchEvent(new Event('input', {{bubbles:true}}));
                inputs[0].dispatchEvent(new Event('change', {{bubbles:true}}));
                inputs[1].value = '{PASSWORD}';
                inputs[1].dispatchEvent(new Event('input', {{bubbles:true}}));
                inputs[1].dispatchEvent(new Event('change', {{bubbles:true}}));
            }}
            var buttons = document.querySelectorAll('.v-button');
            for(var b of buttons) {{
                if(b.textContent.includes('Вход')) {{ b.click(); break; }}
            }}
        }}""")
        await asyncio.sleep(20)
        
        print("[3/7] Навигация...")
        await page.mouse.click(*MENU_DOP_USLUGI)
        await asyncio.sleep(3)
        await page.mouse.click(*MENU_OTCHETNOST)
        await asyncio.sleep(5)
        
        print("[4/7] Открытие отчёта #13...")
        await page.dblclick('text=13. Движение по импорту')
        await asyncio.sleep(5)
        
        print("[5/7] Установка дат...")
        await page.evaluate(f"""() => {{
            var inputs = document.querySelectorAll('.v-datefield-textfield');
            if(inputs.length >= 2) {{
                inputs[0].value = '{start_date} 00:00';
                inputs[0].dispatchEvent(new Event('change', {{bubbles:true}}));
                inputs[1].value = '{end_date} 23:59';
                inputs[1].dispatchEvent(new Event('change', {{bubbles:true}}));
            }}
        }}""")
        await asyncio.sleep(2)
        
        print("[6/7] Генерация отчёта...")
        await page.evaluate("""() => {
            var buttons = document.querySelectorAll('.v-button');
            for(var b of buttons) {
                if(b.textContent.includes('Показать отчет')) {
                    b.click();
                    return true;
                }
            }
            return false;
        }""")
        await asyncio.sleep(20)
        
        screenshot_path = output_file.replace('.xlsx', '.png')
        await page.screenshot(path=screenshot_path, full_page=True)
        print(f"Скриншот: {screenshot_path}")
        
        await browser.close()
        
        print("[7/7] OCR...")
        table_img = preprocess_for_ocr(screenshot_path)
        table_img.save(screenshot_path.replace('.png', '_table.png'))
        
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(table_img, lang='rus+eng', config=custom_config)
        
        # Сохраняем текст для отладки
        with open(screenshot_path.replace('.png', '_ocr.txt'), 'w', encoding='utf-8') as f:
            f.write(text)
        
        print(f"Распознано символов: {len(text)}")
        
        data = parse_ocr_text(text)
        print(f"Найдено {len(data)} контейнеров")
        
        if data:
            create_excel(data, start_date, end_date, output_file)
            return True
        else:
            print("⚠️ OCR не распознал данные")
            return False


def main():
    parser = argparse.ArgumentParser(description='Извлечь отчёт #13 из ДЕЛО ТЕХ с OCR')
    parser.add_argument('--start', required=True, help='Начальная дата (ДД.ММ.ГГГГ)')
    parser.add_argument('--end', required=True, help='Конечная дата (ДД.ММ.ГГГГ)')
    parser.add_argument('--output', default='report_13_ocr.xlsx', help='Имя выходного файла')
    
    args = parser.parse_args()
    
    success = asyncio.run(extract_report(args.start, args.end, args.output))
    if success:
        print(f"\n✅ Готово: {args.output}")
    else:
        print(f"\n⚠️ Проверьте скриншот: {args.output.replace('.xlsx', '.png')}")


if __name__ == '__main__':
    main()
